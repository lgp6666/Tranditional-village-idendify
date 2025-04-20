from django.shortcuts import render
from django.http import HttpResponse
from classifyclass import settings
from .prediction import MyPrediction
import openpyxl


# Create your views here.
def index(request):
    return render(request, 'index.html')


def classifyinfo(request):
    if request.method == "POST":
        f1 = request.FILES['pic1']
        # 用于识别
        fname = '%s/pic/%s' % (settings.MEDIA_ROOT, f1.name)
        with open(fname, 'wb') as pic:
            for c in f1.chunks():
                pic.write(c)
        # 用于显示
        fname1 = './static/img/%s' % f1.name
        with open(fname1, 'wb') as pic:
            for c in f1.chunks():
                pic.write(c)

        preds, real_confidence = MyPrediction(f1.name)

        myExcel = openpyxl.load_workbook('./media/message.xlsx')  # 获取表格文件
        mySheet = myExcel['Sheet1']  # 获取指定的sheet
        finalResult = ['', '', '', '', '']
        if real_confidence < 0:
            finalResult[0] = '无关图片'
        else:
            finalResult[0] = (mySheet.cell(row=preds + 2, column=2)).value
            finalResult[1] = (mySheet.cell(row=preds + 2, column=3)).value
            finalResult[2] = (mySheet.cell(row=preds + 2, column=4)).value
            finalResult[3] = (mySheet.cell(row=preds + 2, column=5)).value
            finalResult[4] = (mySheet.cell(row=preds + 2, column=6)).value

        return render(request, 'info.html',
                      {'finalResult0': finalResult[0], 'finalResult1': finalResult[1],
                       'finalResult2': finalResult[2], 'finalResult3': finalResult[3],
                       'finalResult4': finalResult[4], 'conf': format(real_confidence, ".4f"),
                       'picname': f1.name})
    else:
        return HttpResponse("上传失败！")
